from pathlib import Path
from datetime import datetime
from openpyxl import load_workbook

from app.domain.models.lead import Lead


class AnfragenRepository:
    def __init__(self, file_path: Path, sheet_name: str = None):
        self.file_path = Path(file_path)
        self.sheet_name = sheet_name

    def _get_ws(self):
        wb = load_workbook(self.file_path)
        ws = wb[self.sheet_name] if self.sheet_name else wb.active
        return wb, ws

    def get_headers(self) -> list[str]:
        wb, ws = self._get_ws()
        return [cell.value for cell in ws[1]]

    def get_existing_keys(self) -> set[str]:
        wb, ws = self._get_ws()
        headers = [cell.value for cell in ws[1]]

        required = ["Vorname", "Nachname", "Datum", "Uhrzeit", "Lead-ID", "E-Mail"]
        for col in required:
            if col not in headers:
                raise ValueError(f"Excel must contain column '{col}'")

        nach_index = headers.index("Nachname")
        date_index = headers.index("Datum")
        time_index = headers.index("Uhrzeit")
        lead_index = headers.index("Lead-ID")
        email_index = headers.index("E-Mail")

        existing_keys = set()

        for row in ws.iter_rows(min_row=2, values_only=True):
            lead_val = row[lead_index]
            if lead_val:
                existing_keys.add(str(lead_val).strip().lower())

            email_val = row[email_index]
            if email_val:
                existing_keys.add(str(email_val).strip().lower())

            nach = str(row[nach_index]).strip().lower() if row[nach_index] else ""

            excel_date = row[date_index]
            if isinstance(excel_date, datetime):
                datum = excel_date.date().isoformat()
            elif hasattr(excel_date, "isoformat"):
                datum = excel_date.isoformat()
            else:
                datum = str(excel_date).strip()

            excel_time = row[time_index]
            if isinstance(excel_time, datetime):
                uhrzeit = excel_time.time().strftime("%H:%M:%S")
            elif hasattr(excel_time, "strftime"):
                uhrzeit = excel_time.strftime("%H:%M:%S")
            else:
                uhrzeit = str(excel_time).strip()

            if nach and datum and uhrzeit:
                existing_keys.add(f"sha:{nach}|{datum}|{uhrzeit}".lower())

        return existing_keys

    def append_lead(self, lead: Lead) -> None:
        wb, ws = self._get_ws()
        headers = [cell.value for cell in ws[1]]

        row_map = {
            "Quelle": lead.quelle,
            "Datum": lead.datum,
            "Uhrzeit": lead.uhrzeit,
            "Vorname": lead.vorname,
            "Nachname": lead.nachname,
            "Adresse": lead.adresse,
            "E-Mail": lead.email,
            "Telefon": lead.telefon,
            "Bundesland": lead.bundesland,
            "Bezirk": lead.bezirk,
            "Lead-ID": lead.lead_id,
            "Technologie": lead.technologie,
            "Baustellen-Typ": lead.baustellen_typ,
            "Platzierung": lead.platzierung,
            "Dach": lead.dach,
            "Fläche": lead.flaeche,
            "Besitzverhältnis": lead.besitzverhaeltnis,
            "Zeitraum": lead.zeitraum,
            "Jahresstromverbrauch kWh": lead.jahresstromverbrauch_kwh,
            "Stromspeicher": lead.stromspeicher,
            "Finanzierung": lead.finanzierung,
            "Ergänzende Informationen": lead.ergaenzende_informationen,
        }

        new_row = [row_map.get(header, "") for header in headers]
        ws.append(new_row)
        wb.save(self.file_path)