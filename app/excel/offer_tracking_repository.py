from pathlib import Path
from openpyxl import load_workbook
from openpyxl.styles import Font
from openpyxl.worksheet.hyperlink import Hyperlink

from app.domain.models.offer import Offer


class OfferTrackingRepository:
    def __init__(self, file_path: Path, sheet_name: str | None = None):
        self.file_path = Path(file_path)
        self.sheet_name = sheet_name

    def _get_ws(self):
        wb = load_workbook(self.file_path)
        ws = wb[self.sheet_name] if self.sheet_name else wb.active
        return wb, ws

    def get_headers(self) -> list[str]:
        wb, ws = self._get_ws()
        return [cell.value for cell in ws[1]]

    def get_existing_offer_numbers(self) -> set[str]:
        wb, ws = self._get_ws()
        headers = [cell.value for cell in ws[1]]
        if "AN.-Nr.:" not in headers:
            raise ValueError("Excel must contain column 'AN.-Nr.:'")

        offer_index = headers.index("AN.-Nr.:")
        existing = set()

        for row in ws.iter_rows(min_row=2, values_only=True):
            offer_nr = row[offer_index]
            if offer_nr:
                existing.add(str(offer_nr).strip().lower())

        return existing

    def append_offer(self, offer: Offer) -> None:
        wb, ws = self._get_ws()
        headers = [cell.value for cell in ws[1]]
        print(offer.angebot_pfad_mac)
        row_map = {
            "Name": offer.name,
            "Angebotsdatum": offer.angebotsdatum,
            "angerufen am": offer.angerufen_am,
            "Kontaktperson": offer.kontaktperson,
            "AN.-Nr.:": offer.angebot_nr,
            "Angebot MAC": offer.angebot_pfad_mac,
            "Angebot Windows": offer.angebot_pfad_windows,
            "Bemerkung": offer.bemerkung,
            "urgieren am": offer.urgieren_am,
        }

        new_row = [row_map.get(header, "") for header in headers]
        ws.append(new_row)

        # hyperlink handling for "Angebot Windows"
        if "Angebot Windows" in headers and offer.angebot_pfad_windows:
            angebot_col_windows = headers.index("Angebot Windows") + 1
            row_number = ws.max_row
            cell = ws.cell(row=row_number, column=angebot_col_windows)
            cell.value = "Öffnen"
            cell.hyperlink = offer.angebot_pfad_windows
            cell.font = Font(color="0000FF", underline="single")

        # hyperlink handling for "Angebot MAC"
        if "Angebot MAC" in headers and offer.angebot_pfad_mac:
            angebot_col_mac = headers.index("Angebot MAC") + 1
            row_number = ws.max_row
            cell = ws.cell(row=row_number, column=angebot_col_mac)
            cell.value = "Öffnen"
            cell.hyperlink = Hyperlink(
                ref=cell.coordinate,
                location=None,
                target=offer.angebot_pfad_mac,
                tooltip="PDF Öffnen"
            )
            cell.font = Font(color="0000FF", underline="single")

        wb.save(self.file_path)