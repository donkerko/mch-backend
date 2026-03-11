import os
import re
from datetime import datetime
from openpyxl import load_workbook

# ==========================
# CONFIG
# ==========================

ANGEBOTE_PATH = r"C:\Users\PV\OneDrive\ENKO GmbH\Angebote zu erstellen.xlsx"
ANFRAGEN_PATH = r"C:\Users\PV\OneDrive\ENKO GmbH\Anfragen\Anfragen.xlsx"
TARGET_DIRECTORY = r"C:\Users\PV\OneDrive\ENKO GmbH\Projekte"

NAME_COLUMN = "B"
START_ROW = 2

# ==========================
# HELPERS
# ==========================

def normalize_for_folder(text):
    text = text.strip()

    umlauts = {
        "ä": "ae", "ö": "oe", "ü": "ue",
        "Ä": "Ae", "Ö": "Oe", "Ü": "Ue",
        "ß": "ss"
    }

    for u, r in umlauts.items():
        text = text.replace(u, r)

    return re.sub(r'[<>:"/\\|?*]', '', text)


def normalize_for_compare(text):
    return normalize_for_folder(text).replace(" ", "").lower()


def create_project_structure(project_path):
    structure = [
        "01_Lead",
        os.path.join("02_Datenaufnahme", "Dokumente"),
        os.path.join("02_Datenaufnahme", "Bilder"),
        "03_Planung",
        "04_Angebot",
        "05_Auftrag",
        "06_Rechnung",
        "99_Archiv"
    ]

    for folder in structure:
        os.makedirs(os.path.join(project_path, folder), exist_ok=True)


def project_already_exists(compare_key):
    for existing in os.listdir(TARGET_DIRECTORY):
        if normalize_for_compare(existing.split("_", 1)[-1]) == compare_key:
            return True
    return False


# ==========================
# LOAD ANFRAGEN (DATE SOURCE)
# ==========================

print("📖 Lade Anfragen.xlsx...")

wb_anfragen = load_workbook(ANFRAGEN_PATH)
ws_anfragen = wb_anfragen.active

date_lookup = {}

for row in range(2, ws_anfragen.max_row + 1):

    date_value = ws_anfragen[f"C{row}"].value
    first = ws_anfragen[f"E{row}"].value
    last = ws_anfragen[f"F{row}"].value

    if not date_value or not first or not last:
        continue

    # --- DATE HANDLING ---
    if isinstance(date_value, datetime):
        formatted_date = date_value.strftime("%Y-%m-%d")
    else:
        try:
            parsed_date = datetime.strptime(str(date_value), "%d/%m/%Y")
            formatted_date = parsed_date.strftime("%Y-%m-%d")
        except ValueError:
            print(f"⚠ Ungültiges Datum in Zeile {row}")
            continue

    key1 = normalize_for_compare(f"{first}{last}")
    key2 = normalize_for_compare(f"{last}{first}")

    date_lookup[key1] = formatted_date
    date_lookup[key2] = formatted_date

print("✅ Anfragen geladen.\n")


# ==========================
# PROCESS ANGEBOTE LIST
# ==========================

wb_angebote = load_workbook(ANGEBOTE_PATH)
ws_angebote = wb_angebote.active

for row in range(START_ROW, ws_angebote.max_row + 1):

    name_cell = ws_angebote[f"{NAME_COLUMN}{row}"].value

    if not name_cell:
        continue

    original_name = str(name_cell).strip()
    compare_key = normalize_for_compare(original_name)

    if compare_key not in date_lookup:
        print(f"❌ Kein Datum gefunden für: {original_name}")
        continue

    if project_already_exists(compare_key):
        print(f"⚠ Projekt existiert bereits: {original_name}")
        continue

    project_date = date_lookup[compare_key]
    normalized_name = normalize_for_folder(original_name)

    project_folder_name = f"{project_date}_{normalized_name}"
    project_path = os.path.join(TARGET_DIRECTORY, project_folder_name)

    create_project_structure(project_path)

    print(f"✅ Erstellt: {project_folder_name}")

print("\n🎯 Fertig.")