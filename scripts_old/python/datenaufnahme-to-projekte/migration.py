import os
import shutil
import zipfile
import re
from openpyxl import load_workbook
from datetime import datetime

# ==========================
# CONFIG
# ==========================

DATENAUFNAHMEN_PATH = r"C:\Users\PV\OneDrive\ENKO GmbH\Datenaufnahmen"
PROJEKTE_PATH = r"C:\Users\PV\OneDrive\ENKO GmbH\Projekte"
ANFRAGEN_PATH = r"C:\Users\PV\OneDrive\ENKO GmbH\Anfragen\Anfragen.xlsx"

# ==========================
# HELPERS
# ==========================

def normalize_for_compare(text):
    if not text:
        return ""

    text = text.strip()

    umlauts = {
        "ä": "ae", "ö": "oe", "ü": "ue",
        "Ä": "Ae", "Ö": "Oe", "Ü": "Ue",
        "ß": "ss"
    }

    for u, r in umlauts.items():
        text = text.replace(u, r)

    text = re.sub(r'[<>:"/\\|?*]', '', text)
    return text.replace(" ", "").lower()


def normalize_for_folder(text):
    text = text.strip()

    umlauts = {
        "ä": "ae", "ö": "oe", "ü": "ue",
        "Ä": "Ae", "Ö": "Oe", "Ü": "Ue",
        "ß": "ss"
    }

    for u, r in umlauts.items():
        text = text.replace(u, r)

    return re.sub(r'[<>:"/\\|?*]', '', text)


def create_project_structure(project_path):
    structure = [
        "01_Lead",
        os.path.join("02_Datenaufnahme", "Dokumente"),
        os.path.join("02_Datenaufnahme", "Bilder"),
        "03_Planung",
        "04_Angebot",
        "05_Auftrag",
        "06_Rechnung",
        "99_Archiv"
    ]

    for folder in structure:
        os.makedirs(os.path.join(project_path, folder), exist_ok=True)


def extract_images_from_docx(docx_path, image_target_folder):
    try:
        with zipfile.ZipFile(docx_path, 'r') as zip_ref:
            for file in zip_ref.namelist():
                if file.startswith("word/media/"):
                    zip_ref.extract(file, image_target_folder)

        media_path = os.path.join(image_target_folder, "word", "media")

        if os.path.exists(media_path):
            for img in os.listdir(media_path):
                shutil.move(
                    os.path.join(media_path, img),
                    os.path.join(image_target_folder, img)
                )

            shutil.rmtree(os.path.join(image_target_folder, "word"))

    except Exception as e:
        print(f"⚠ Fehler bei Bild-Extraktion: {e}")


# ==========================
# LOAD EXCEL DATA
# ==========================

print("📖 Lade Anfragen.xlsx...")

wb = load_workbook(ANFRAGEN_PATH)
ws = wb.active

date_lookup = {}

for row in range(2, ws.max_row + 1):

    date_value = ws[f"C{row}"].value
    first = ws[f"E{row}"].value
    last = ws[f"F{row}"].value

    if not date_value or not first or not last:
        continue

    # --- DATE HANDLING ---
    if isinstance(date_value, datetime):
        formatted_date = date_value.strftime("%Y-%m-%d")
    else:
        try:
            parsed_date = datetime.strptime(str(date_value), "%d/%m/%Y")
            formatted_date = parsed_date.strftime("%Y-%m-%d")
        except ValueError:
            print(f"⚠ Ungültiges Datum in Zeile {row}: {date_value}")
            continue

    key1 = normalize_for_compare(f"{first}{last}")
    key2 = normalize_for_compare(f"{last}{first}")

    date_lookup[key1] = formatted_date
    date_lookup[key2] = formatted_date

print("✅ Excel geladen.\n")

# ==========================
# MAIN MIGRATION
# ==========================

for folder_name in os.listdir(DATENAUFNAHMEN_PATH):

    source_folder = os.path.join(DATENAUFNAHMEN_PATH, folder_name)

    if not os.path.isdir(source_folder):
        continue

    compare_key = normalize_for_compare(folder_name)

    if compare_key not in date_lookup:
        print(f"❌ Kein Datum gefunden für: {folder_name}")
        continue

    project_date = date_lookup[compare_key]
    normalized_folder_name = normalize_for_folder(folder_name)

    project_folder_name = f"{project_date}_{normalized_folder_name}"
    project_path = os.path.join(PROJEKTE_PATH, project_folder_name)

    print(f"\n🔄 Migration: {folder_name} → {project_folder_name}")

    create_project_structure(project_path)

    for file in os.listdir(source_folder):

        source_file = os.path.join(source_folder, file)

        if not os.path.isfile(source_file):
            continue

        lower_file = file.lower()

        if lower_file.endswith(".docx"):
            target_doc_folder = os.path.join(project_path, "02_Datenaufnahme", "Dokumente")
            target_doc_path = os.path.join(target_doc_folder, file)

            shutil.copy2(source_file, target_doc_path)

            extract_images_from_docx(
                target_doc_path,
                os.path.join(project_path, "02_Datenaufnahme", "Bilder")
            )

        elif lower_file.endswith((".jpg", ".jpeg", ".png")):
            target_img_folder = os.path.join(project_path, "02_Datenaufnahme", "Bilder")
            shutil.copy2(source_file, target_img_folder)

        else:
            target_doc_folder = os.path.join(project_path, "02_Datenaufnahme", "Dokumente")
            shutil.copy2(source_file, target_doc_folder)

    print("✅ Fertig.")

print("\n🎯 Migration abgeschlossen.")