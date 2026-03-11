import os
import shutil
import re
import pdfplumber
import logging
from datetime import datetime
import win32com.client
from openpyxl import load_workbook

# ==========================================
# CONFIG
# ==========================================

SCRIPT_NAME = "sicherheitssysteme_pipeline"

MAILBOX_NAME = "huetter@mchvertrieb.at"
SOURCE_FOLDER = "Sicherheitssysteme"
PROCESSED_FOLDER = "Gespeichert"

INCOMING_DIR = r"C:\Users\PV\OneDrive\Sicherheitssysteme Vöcklabruck GmbH\Angebote"
PROCESSED_DIR = r"C:\Offers\Processed"

EXCEL_FILE = r"C:\Users\PV\OneDrive\Sicherheitssysteme Vöcklabruck GmbH\SYSS Angebote - Kontaktliste.xlsx

LOG_DIR = r"C:\Users\PV\OneDrive\backend\logs\sicherheitssysteme"

# ==========================================
# LOGGING
# ==========================================

os.makedirs(LOG_DIR, exist_ok=True)

log_file = os.path.join(
    LOG_DIR,
    f"{SCRIPT_NAME}_{datetime.now().date()}.log"
)

logging.basicConfig(
    filename=log_file,
    level=logging.INFO,
    format="%(asctime)s | %(message)s"
)

def log(customer, action):
    logging.info(f"{SCRIPT_NAME} | {customer} | {action}")

# ==========================================
# OUTLOOK
# ==========================================

def get_outlook_folders():

    outlook = win32com.client.Dispatch("Outlook.Application")
    namespace = outlook.GetNamespace("MAPI")

    mailbox = namespace.Folders[MAILBOX_NAME]

    source_folder = mailbox.Folders[SOURCE_FOLDER]
    target_folder = source_folder.Folders[PROCESSED_FOLDER]

    return source_folder, target_folder

# ==========================================
# SAVE PDF ATTACHMENTS
# ==========================================

def process_outlook():

    source_folder, target_folder = get_outlook_folders()

    items = source_folder.Items
    items.Sort("[ReceivedTime]", True)

    pdf_files = []

    for mail in list(items):

        try:

            if mail.Class != 43:
                continue

            saved_any_pdf = False

            for attachment in mail.Attachments:

                filename = attachment.FileName

                if not filename.lower().endswith(".pdf"):
                    continue

                save_path = os.path.join(INCOMING_DIR, filename)

                counter = 1

                while os.path.exists(save_path):
                    name, ext = os.path.splitext(filename)
                    save_path = os.path.join(
                        INCOMING_DIR,
                        f"{name}_{counter}{ext}"
                    )
                    counter += 1

                attachment.SaveAsFile(save_path)

                pdf_files.append(save_path)

                saved_any_pdf = True

                log("unknown", f"PDF saved: {save_path}")

            if saved_any_pdf:

                mail.Move(target_folder)

                log("unknown", "Mail moved to Sicherheitssysteme/Gespeichert")

        except Exception as e:

            log("unknown", f"Outlook error: {e}")

    return pdf_files

# ==========================================
# PDF PARSER
# ==========================================

def parse_pdf(pdf_path):

    text = ""

    with pdfplumber.open(pdf_path) as pdf:
        for page in pdf.pages:
            txt = page.extract_text()
            if txt:
                text += txt + "\n"

    lines = text.split("\n")

    name = ""
    contact = ""

    if len(lines) > 0:
        name = lines[0].strip()

    if len(lines) > 1:
        contact = lines[1].strip()

    date_match = re.search(r"\d{1,2}\.\d{1,2}\.\d{4}", text)
    angebot_datum = date_match.group() if date_match else ""

    offer_match = re.search(r"Angebot:\s*(\S+)", text)
    offer_number = offer_match.group(1) if offer_match else ""

    return {
        "name": name,
        "kontaktperson": contact,
        "datum": angebot_datum,
        "offer": offer_number
    }

# ==========================================
# MOVE PDF
# ==========================================

def move_pdf(file_path):

    filename = os.path.basename(file_path)

    new_path = os.path.join(PROCESSED_DIR, filename)

    counter = 1

    while os.path.exists(new_path):

        name, ext = os.path.splitext(filename)

        new_path = os.path.join(
            PROCESSED_DIR,
            f"{name}_{counter}{ext}"
        )

        counter += 1

    shutil.move(file_path, new_path)

    return new_path

# ==========================================
# EXCEL UPDATE
# ==========================================

def update_excel(data, pdf_path):

    wb = load_workbook(EXCEL_FILE)
    ws = wb.active

    next_row = ws.max_row + 1

    ws.cell(next_row, 1).value = data["name"]
    ws.cell(next_row, 2).value = data["datum"]
    ws.cell(next_row, 3).value = ""  # angerufen am
    ws.cell(next_row, 4).value = data["kontaktperson"]
    ws.cell(next_row, 5).value = data["offer"]
    ws.cell(next_row, 6).value = pdf_path
    ws.cell(next_row, 7).value = ""
    ws.cell(next_row, 8).value = ""

    wb.save(EXCEL_FILE)

    log(data["name"], f"Excel updated {data['offer']}")

# ==========================================
# MAIN PIPELINE
# ==========================================

def main():

    os.makedirs(INCOMING_DIR, exist_ok=True)
    os.makedirs(PROCESSED_DIR, exist_ok=True)

    log("system", "Pipeline started")

    pdf_files = process_outlook()

    for pdf in pdf_files:

        try:

            data = parse_pdf(pdf)

            new_pdf_path = move_pdf(pdf)

            update_excel(data, new_pdf_path)

            log(data["name"], "Offer processed")

        except Exception as e:

            log("unknown", f"Processing error: {e}")

    log("system", "Pipeline finished")


if __name__ == "__main__":
    main()