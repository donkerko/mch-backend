import os
from openpyxl import load_workbook

# Pfade
folder_path = r"C:\Users\PV\OneDrive\ENKO GmbH\Datenaufnahmen"
excel_path = r"C:\Users\PV\OneDrive\ENKO GmbH\Projektübersicht.xlsx"

# Excel öffnen
wb = load_workbook(excel_path)
ws = wb.active

# Alle existierenden Kundennamen in Spalte B sammeln
existing_customers = set()
for cell in ws['B']:
    if cell.value:
        existing_customers.add(cell.value.strip())

# Unterordner im Datenaufnahmen-Ordner auflisten
subfolders = [f for f in os.listdir(folder_path) if os.path.isdir(os.path.join(folder_path, f))]

# Neue Kunden hinzufügen
new_entries = 0
for customer in subfolders:
    if customer not in existing_customers:
        # Nächste freie Zeile ermitteln
        next_row = ws.max_row + 1
        ws[f'B{next_row}'] = customer
        new_entries += 1

# Excel speichern
wb.save(excel_path)

print(f"{new_entries} neue Kunden wurden hinzugefügt.")
