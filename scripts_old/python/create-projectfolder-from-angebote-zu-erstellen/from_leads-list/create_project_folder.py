import os
import re
from datetime import datetime, date
from openpyxl import load_workbook

# =========================================================
# CONFIG
# =========================================================

SCRIPT_NAME = "create_projectfolders_from_leads.py"

LEADS_PATH = r"C:\Users\PV\OneDrive\ENKO GmbH\Leads\Leads.xlsx"
PROJECTS_PATH = r"C:\Users\PV\OneDrive\ENKO GmbH\Projekte"

GLOBAL_LOG_DIR = r"C:\Users\PV\OneDrive\backend\logs"
GLOBAL_LOG_FILE = os.path.join(GLOBAL_LOG_DIR, "create_projectfolders_from_leads.log")

# =========================================================
# LOGGING
# =========================================================

def log_global(customer, action):
    os.makedirs(GLOBAL_LOG_DIR, exist_ok=True)
    timestamp = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    line = f"{timestamp} | {SCRIPT_NAME} | {customer} | {action}\n"

    with open(GLOBAL_LOG_FILE, "a", encoding="utf-8") as f:
        f.write(line)


def log_local(project_path, action):
    os.makedirs(project_path, exist_ok=True)
    log_file = os.path.join(project_path, "log.txt")

    timestamp = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    line = f"{timestamp} | {SCRIPT_NAME} | {action}\n"

    with open(log_file, "a", encoding="utf-8") as f:
        f.write(line)

# =========================================================
# HELPERS
# =========================================================

TITLES = {
    "mag", "mag.", "dr", "dr.", "bsc", "bsc.", "msc", "msc.",
    "dipl", "dipl.", "ing", "ing.", "mba", "mba.", "phd", "phd.",
    "prof", "prof.", "ba", "ba.", "llm", "llm."
}


def remove_titles(text):
    if not text:
        return ""

    parts = str(text).strip().split()
    cleaned = []

    for part in parts:
        normalized = part.strip().lower()
        if normalized not in TITLES:
            cleaned.append(part)

    return " ".join(cleaned).strip()


def normalize_for_folder(text):
    if not text:
        return ""

    text = remove_titles(str(text).strip())

    replacements = {
        "ä": "ae", "ö": "oe", "ü": "ue",
        "Ä": "Ae", "Ö": "Oe", "Ü": "Ue",
        "ß": "ss"
    }

    for old, new in replacements.items():
        text = text.replace(old, new)

    # remove invalid Windows characters
    text = re.sub(r'[<>:"/\\|?*]', "", text)

    # collapse spaces
    text = re.sub(r"\s+", " ", text).strip()

    return text


def normalize_for_compare(text):
    text = normalize_for_folder(text)
    return text.replace(" ", "").lower()


def split_name(full_name):
    full_name = remove_titles(full_name).strip()

    if not full_name:
        return "", ""

    parts = full_name.split()

    if len(parts) == 1:
        return parts[0], ""

    vorname = parts[0]
    nachname = " ".join(parts[1:])
    return vorname, nachname


def build_name_variants(vorname, nachname):
    variants = []

    if vorname and nachname:
        variants.append(f"{vorname} {nachname}")
        variants.append(f"{nachname} {vorname}")
    elif vorname:
        variants.append(vorname)
    elif nachname:
        variants.append(nachname)

    # normalized unique variants
    seen = set()
    unique_variants = []

    for variant in variants:
        key = normalize_for_compare(variant)
        if key and key not in seen:
            seen.add(key)
            unique_variants.append(variant)

    return unique_variants


def parse_excel_date(value):
    if value is None or value == "":
        return None

    if isinstance(value, datetime):
        return value.date()

    if isinstance(value, date):
        return value

    text = str(value).strip()

    date_formats = [
        "%Y-%m-%d",
        "%d.%m.%Y",
        "%d/%m/%Y",
        "%Y/%m/%d",
        "%d-%m-%Y",
    ]

    for fmt in date_formats:
        try:
            return datetime.strptime(text, fmt).date()
        except ValueError:
            continue

    return None


def create_project_structure(project_path):
    structure = [
        "01_Lead",
        os.path.join("02_Datenaufnahme", "Dokumente"),
        os.path.join("02_Datenaufnahme", "Bilder"),
        "03_Planung",
        "04_Angebot",
        "05_Auftrag",
        "06_Rechnung",
        "99_Archiv"
    ]

    for folder in structure:
        os.makedirs(os.path.join(project_path, folder), exist_ok=True)


def get_existing_project_folders():
    existing = []

    if not os.path.exists(PROJECTS_PATH):
        os.makedirs(PROJECTS_PATH, exist_ok=True)
        return existing

    for folder_name in os.listdir(PROJECTS_PATH):
        folder_path = os.path.join(PROJECTS_PATH, folder_name)

        if not os.path.isdir(folder_path):
            continue

        # expected format: YYYY-MM-DD_Name
        name_part = folder_name
        if "_" in folder_name:
            parts = folder_name.split("_", 1)
            if len(parts) == 2:
                name_part = parts[1]

        existing.append({
            "folder_name": folder_name,
            "folder_path": folder_path,
            "compare_key": normalize_for_compare(name_part)
        })

    return existing


def project_already_exists(name_variants, existing_projects):
    variant_keys = {normalize_for_compare(v) for v in name_variants if v}

    for project in existing_projects:
        if project["compare_key"] in variant_keys:
            return project

    return None


def find_column_index(headers, possible_names):
    normalized_headers = [str(h).strip().lower() if h is not None else "" for h in headers]

    for name in possible_names:
        name_lower = name.strip().lower()
        if name_lower in normalized_headers:
            return normalized_headers.index(name_lower)

    return None

# =========================================================
# MAIN
# =========================================================

def main():
    print("Lade Leads.xlsx ...")

    wb = load_workbook(LEADS_PATH)
    ws = wb.active

    headers = [cell.value for cell in ws[1]]

    vorname_idx = find_column_index(headers, ["Vorname", "First Name"])
    nachname_idx = find_column_index(headers, ["Nachname", "Last Name", "Surname"])
    name_idx = find_column_index(headers, ["Name", "Kunde", "Kundenname", "Customer"])
    datum_idx = find_column_index(headers, ["Datum", "Erstellt am", "Lead-Datum", "Date"])

    if (vorname_idx is None or nachname_idx is None) and name_idx is None:
        raise ValueError(
            "Keine passende Namensspalte gefunden. "
            "Benötigt entweder 'Vorname' + 'Nachname' oder eine einzelne Spalte wie 'Name'/'Kunde'."
        )

    existing_projects = get_existing_project_folders()

    created_count = 0
    skipped_existing_count = 0
    skipped_no_name_count = 0

    for row_number, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        # -------------------------
        # Name lesen
        # -------------------------
        vorname = ""
        nachname = ""

        if vorname_idx is not None and nachname_idx is not None:
            vorname = str(row[vorname_idx]).strip() if row[vorname_idx] else ""
            nachname = str(row[nachname_idx]).strip() if row[nachname_idx] else ""

        elif name_idx is not None and row[name_idx]:
            full_name = str(row[name_idx]).strip()
            vorname, nachname = split_name(full_name)

        vorname = remove_titles(vorname)
        nachname = remove_titles(nachname)

        if not vorname and not nachname:
            print(f"Zeile {row_number}: übersprungen, kein Name")
            skipped_no_name_count += 1
            log_global("UNBEKANNT", f"Zeile {row_number} übersprungen - kein Name")
            continue

        display_name = f"{vorname} {nachname}".strip()
        if not display_name:
            display_name = vorname or nachname

        name_variants = build_name_variants(vorname, nachname)

        # -------------------------
        # Existenz prüfen
        # both firstname lastname and lastname firstname
        # -------------------------
        existing_project = project_already_exists(name_variants, existing_projects)

        if existing_project:
            print(f"Bereits vorhanden: {display_name} -> {existing_project['folder_name']}")
            skipped_existing_count += 1
            log_global(display_name, f"Projektordner existiert bereits: {existing_project['folder_name']}")
            continue

        # -------------------------
        # Datum lesen
        # -------------------------
        project_date = None

        if datum_idx is not None:
            project_date = parse_excel_date(row[datum_idx])

        if project_date is None:
            project_date = datetime.today().date()
            log_global(display_name, f"Kein gültiges Datum gefunden - heutiges Datum verwendet: {project_date.isoformat()}")

        formatted_date = project_date.strftime("%Y-%m-%d")

        # Standard folder naming: Vorname Nachname
        base_name = f"{vorname} {nachname}".strip()
        folder_name = f"{formatted_date}_{normalize_for_folder(base_name)}"
        project_path = os.path.join(PROJECTS_PATH, folder_name)

        # final safety check
        if os.path.exists(project_path):
            print(f"Ordner existiert schon physisch: {folder_name}")
            log_global(display_name, f"Ordner bereits vorhanden (physisch): {folder_name}")
            continue

        # -------------------------
        # Ordner erstellen
        # -------------------------
        create_project_structure(project_path)

        action = f"Projektordner erstellt: {folder_name}"
        log_global(display_name, action)
        log_local(project_path, action)

        print(f"Erstellt: {folder_name}")

        existing_projects.append({
            "folder_name": folder_name,
            "folder_path": project_path,
            "compare_key": normalize_for_compare(base_name)
        })

        # also store reversed name variant in memory via extra entry
        if vorname and nachname:
            reversed_name = f"{nachname} {vorname}"
            existing_projects.append({
                "folder_name": folder_name,
                "folder_path": project_path,
                "compare_key": normalize_for_compare(reversed_name)
            })

        created_count += 1

    print("\nFertig.")
    print(f"Erstellt: {created_count}")
    print(f"Bereits vorhanden: {skipped_existing_count}")
    print(f"Ohne Namen übersprungen: {skipped_no_name_count}")


if __name__ == "__main__":
    main()