from openpyxl import load_workbook

# Paths to the Excel files
FIRST_FILE = "C:/Users/PV/OneDrive/ENKO GmbH/Projektübersicht.xlsx"   # File where we insert data
SECOND_FILE = "C:/Users/PV/OneDrive/ENKO GmbH/260303_kopie_Leads-Marcel-Huetter.xlsx" # File to look up values

# Open the first workbook and get its sheet name
wb1 = load_workbook(FIRST_FILE)
sheet_name = wb1.sheetnames[0]  # assuming you want the first sheet
ws1 = wb1[sheet_name]

# Open the second workbook
wb2 = load_workbook(SECOND_FILE)
ws2 = wb2.active  # assuming lookup is in the active sheet

# Settings
NAME_COLUMN_WS1 = "B"       # column in first sheet with names to look up
LOOKUP_COLUMN_WS2 = "B"     # column in second sheet where names are listed
EXTRACT_COLUMNS_WS2 = ("C", "D", "E")  # columns to extract from second sheet
INSERT_START_COLUMN_WS1 = "C"     # first column where extracted data will go in sheet 1

# Convert column letter to index (0-based)
def col_to_index(col):
    return ord(col.upper()) - 65

lookup_index_ws2 = col_to_index(LOOKUP_COLUMN_WS2)
extract_indices_ws2 = [col_to_index(c) for c in EXTRACT_COLUMNS_WS2]
insert_index_ws1 = col_to_index(INSERT_START_COLUMN_WS1)

# Process each name in first sheet
for row in ws1.iter_rows(min_row=2):  # skip header
    name = row[col_to_index(NAME_COLUMN_WS1)].value
    if not name:
        continue

    # Find name in second sheet
    found = False
    for lookup_row in ws2.iter_rows(min_row=2, values_only=False):
        if lookup_row[lookup_index_ws2].value == name:
            # Extract the two values
            extracted_values = [lookup_row[i].value for i in extract_indices_ws2]

            # Insert into first sheet
            for offset, value in enumerate(extracted_values):
                row[insert_index_ws1 + offset].value = value

            found = True
            break

    if not found:
        print(f"Name '{name}' not found in {SECOND_FILE}")

# Save updated workbook
wb1.save(FIRST_FILE)
print("All data updated successfully!")