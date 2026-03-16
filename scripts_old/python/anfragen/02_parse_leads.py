import os
import pdfplumber
import re
import email
import chardet
import extract_msg
import hashlib
from email import policy
from datetime import datetime
from openpyxl import load_workbook

# ============================================================
# ---------------------- Settings ----------------------------
# ============================================================
FOLDER_PATH_VOLTALUX = r"C:\Users\PV\OneDrive\ENKO GmbH\Leads\Voltalux"
FOLDER_PATH_PHOTOVOLTAIKANLAGEAT = r"C:\Users\PV\OneDrive\ENKO GmbH\Leads\photovoltaikAT"
FOLDER_PATH_PVALARM = r"C:\Users\PV\OneDrive\ENKO GmbH\Leads\PVALARM"

EXCEL_PATH = r"C:\Users\PV\OneDrive\ENKO GmbH\Leads\Leads.xlsx"

# Starting question in MSG files
START_QUESTION = "Welche Kategorie trifft auf Sie zu?"

# Titles
TITLES = {
    "mag", "dr", "bsc", "msc", "dipl", "ing",
    "mba", "phd", "prof", "ba", "llm"
}

# ============================================================
# ------------------ HELPER FUNCTIONS ------------------------
# ============================================================

def normalize(text):
    return str(text).strip().lower() if text else ""

def get_question_value(lines, question):
    for i, line in enumerate(lines):
        if line.startswith(question):
            
            # Fall 1: Antwort in derselben Zeile (nach ?)
            if "?" in line:
                parts = line.split("?", 1)
                if parts[1].strip():
                    return parts[1].strip()
            else:
                return lines[i + 1].strip()
            
            return ""
    return ""

def extract_datetime_from_text(text):
    """Extract date/time from  MSG body."""
    match_msg = re.search(r"Empfangen\s+(\d{4}-\d{2}-\d{2})\s+kl\.\s+(\d{2}:\d{2})", text)
    if match_msg:
        dt = datetime.strptime(match_msg.group(1) + " " + match_msg.group(2), "%Y-%m-%d %H:%M")
        return dt.date(), dt.time()
    return "", ""

def extract_datetime_from_zapier(text):
    """Extract date/time from Zapier Anfragealarm body."""
    match = re.search(r"Zeit:\s*(\d{4}-\d{2}-\d{2})T(\d{2}:\d{2}:\d{2})", text)
    if match:
        dt = datetime.strptime(match.group(1) + " " + match.group(2),"%Y-%m-%d %H:%M:%S")
        return dt.date(), dt.time()
    return "", ""

def clean_email(raw_email):
    """Extract the actual email from a string like 'mailto:email <email>'"""
    if not raw_email:
        return ""
    # Find the first email address in the string
    match = re.search(r'[\w\.-]+@[\w\.-]+', raw_email)
    if match:
        return match.group(0).strip()
    return raw_email.strip()  # fallback

def extract_kwh(text):
    if not text:
        return ""
    
    match = re.search(r"(\d[\d\.]*)\s*kwh", text.lower())
    if not match:
        return ""
    
    value = match.group(1).replace(".", "")
    return int(value)

def clean_name(full_name):
    if not full_name:
        return "", ""

    # Split nach Leerzeichen
    parts = full_name.split()

    cleaned_parts = []

    for part in parts:
        # Punkte entfernen + lowercase vergleichen
        normalized = re.sub(r"\.", "", part).lower()

        if normalized not in TITLES:
            cleaned_parts.append(part)

    if not cleaned_parts:
        return "", ""

    if len(cleaned_parts) == 1:
        return "", cleaned_parts[0]

    return cleaned_parts[0], " ".join(cleaned_parts[1:])


# ============================================================
# ----------- VOLTALUX PDF EXTRACTION ------------------------
# ============================================================

def extract_datetime_pdf(full_text):
    match = re.search(r"Am\s+(\d{2}\.\d{2}\.\d{4})\s+um\s+(\d{2}:\d{2})", full_text)
    if match:
        dt = datetime.strptime(match.group(1) + " " + match.group(2), "%d.%m.%Y %H:%M")
        return dt.date(), dt.time()
    return "", ""

def extract_data_from_pdf(full_text):
    lines = [line.strip() for line in full_text.splitlines() if line.strip()]
    date, time = extract_datetime_pdf(full_text)

    def get_colon_value(label):
        for line in lines:
            if line.startswith(label + ":"):
                return line.split(":", 1)[1].strip()
        return ""

    return {
        "Quelle": "Voltalux",
        "Datum": date,
        "Uhrzeit": time,
        "Vorname": get_colon_value("Vorname"),
        "Nachname": get_colon_value("Nachname"),
        "Adresse": get_colon_value("Adresse"),
        "E-Mail": get_colon_value("E-Mail"),
        "Telefon": get_colon_value("Telefon"),
        "Bundesland": get_colon_value("Bundesland"),
        "Bezirk": get_colon_value("Bezirk"),
        "Lead-ID": get_colon_value("Lead ID"),
        "Baustellen-Typ": get_question_value(lines, "Wo soll Ihre Photovoltaik Anlage installiert werden?"),
        "Platzierung": get_question_value(lines, "Auf welcher Fläche soll die Photovoltaik Anlage installiert werden?"),
        "Besitzverhältnis": get_question_value(lines, "Sind Sie Eigentümer*in der Immobilie?"),
        "Zeitraum": get_question_value(lines, "Wann soll die Anlage installiert werden?"),
        "Jahresstromverbrauch kWh": extract_kwh(get_question_value(lines, "Wie hoch ist ihr durchschnittlicher Jahresstromverbrauch?")),
        "Stromspeicher": "Ja" if "Mit Stromspeicher" in get_question_value(lines, "Möchten Sie die Photovoltaik Anlage mit einem Stromspeicher ergänzen?") else "Nein",
        "Finanzierung": get_question_value(lines, "Wie wird die Anschaffung der Anlage finanziert?"),
        "Ergänzende Informationen": get_question_value(lines, "Zusätzliche Anmerkungen"),
    }

# ============================================================
# ----------- PHOTOVOLTAIKANLAGE.AT MSG EXTRACTION ----------
# ============================================================

def extract_data_from_msg_photovoltaikanlageat(msg_path):
    """Extract fields from Photovoltaikanlage.at MSG file."""
    msg = extract_msg.Message(msg_path)
    body = msg.body
    lines = [line.strip() for line in body.splitlines() if line.strip()]

    # Find start question
    #try:
    #    start_index = lines.index(START_QUESTION)
    #except ValueError:
    #    print(f"Skipping {msg_path}, no start question found")
    #    return None
    start_index = None

    for i, line in enumerate(lines):
        if START_QUESTION in line:
            start_index = i
            break

    if start_index is None:
        print(f"Skipping {msg_path}, no start question found")
        return None

    data = {}

    # detect tab-based format
    if any("\t" in line for line in lines):

        for line in lines:
            if "\t" in line:
                key, value = line.split("\t", 1)

                # stop at footer
                if "© Nettbureau" in key:
                    break

                data[key.strip()] = value.strip()

    else:
        i = start_index
        while i < len(lines) - 1:
            question = lines[i]
            answer = lines[i + 1]
            if "© Nettbureau" in question or "Kontaktieren Sie" in question:
                break
            data[question] = answer
            i += 2

    # Extract name
    full_name = (
        data.get("Ihr Name", "").strip()
        or data.get("Kontaktperson", "").strip()
    )

    vorname, nachname = clean_name(full_name)

    # Extract date/time
    datum, uhrzeit = extract_datetime_from_text(body)

    # Verbrauch
    v = data.get("Wie hoch ist Ihr Energieverbrauch pro Jahr ungefähr?", "").strip()
    verbrauch = int(v) * 1000 if v and int(v) < 100 else int(v) if v else ""



    # Build return dict
    return {
        "Quelle": "Photovoltaikanlage.at",
        "Datum": datum,
        "Uhrzeit": uhrzeit,
        "Vorname": vorname,
        "Nachname": nachname,
        "Adresse": f"{data.get('Adresse','')} {data.get('Hausnr.','')}, {data.get('Postleitzahl','')} {data.get('Ortschaft','')}".strip(),
        "E-Mail": clean_email(data.get("E-Mail-Addresse", "")),
        "Telefon": data.get("Mobile Rufnummer", ""),
        "Bundesland": data.get("Region",""),
        "Bezirk": data.get("Bezirk",""),
        "Lead-ID": "",
        "Technologie": data.get("Welche Technologie interessiert Sie?",""),
        "Besitzverhältnis": data.get("Welche Kategorie trifft auf Sie zu?",""),
        "Platzierung": data.get("Wo soll das System montiert werden?",""),
        "Dach": data.get("Welche Dachform hat das Haus?",""),
        "Fläche": data.get("Wie groß ist die Fläche ungefähr, auf der die Anlage installiert werden soll?",""),
        "Baustellen-Typ": data.get("Wählen Sie den Gebäudetyp aus.",""),
        "Zeitraum": data.get("Wann möchten Sie Ihre Anlage installieren?",""),
        "Jahresstromverbrauch kWh": verbrauch,
        "Überschüssige Energie verkaufen": data.get("Beabsichtigen Sie, überschüssige Energie zu verkaufen?",""),
        "Stromspeicher": "Ja" if "Stromspeicher" in data.get("Sind Sie an den folgenden Optionen interessiert?","") else "Nein",
        "Finanzierung": "Leasing/Mieten" if "Leasing / Mieten" in data.get("Sind Sie an den folgenden Optionen interessiert?","") else "",
        "E-Ladestation": "Ja" if "E-Ladestation" in data.get("Sind Sie an den folgenden Optionen interessiert?","") else "Nein",
        "Notstrom": "Ja" if "Notstrom" in data.get("Ergänzende Informationen","") else "Nein",
        "Sonstige Optionen": data.get("Sind Sie an den folgenden Optionen interessiert?",""),
        "Ergänzende Informationen": data.get("Ergänzende Informationen",""),
    }

# ============================================================
# ----------- PVALARM MSG EXTRACTION ----------
# ============================================================

def extract_data_from_msg_pvalarm(msg_path):
    """Extract fields from PVALARM MSG file."""
    msg = extract_msg.Message(msg_path)
    body = msg.body
    lines = [line.strip() for line in body.splitlines() if line.strip()]
    #print("Answer" + lines.get("Wie konkret ist dein Projekt aktuell?",""))
    data = {}

    for index, line in enumerate(lines):
        # Match "Key: Value"
        match = re.match(r"([^:]+):\s*(.+)", line)
        
        if match:
            key = match.group(1).strip()
            value = match.group(2).strip()
            data[key] = value
        
        # Special case: "Thema PV"
        elif line.startswith("Thema"):
            parts = line.split(" ", 1)
            if len(parts) == 2:
                data["Thema"] = parts[1].strip()

        elif index + 1 < len(lines):
            # Detect question lines without colon
            if line.endswith("?"):
                data[line.strip()] = lines[index + 1].strip()


    # Extract name
    full_name = data.get("Name", "")
    vorname = full_name.split()[0] if full_name else ""
    nachname = " ".join(full_name.split()[1:]) if len(full_name.split()) > 1 else ""

    # Extract date/time
    datum, uhrzeit = extract_datetime_from_zapier(body)

    # Build return dict
    return {
        "Quelle": "PVALARM",
        "Datum": datum,
        "Uhrzeit": uhrzeit,
        "Vorname": vorname,
        "Nachname": nachname,
        "Telefon": data.get("Tel", ""),
        "E-Mail": clean_email(data.get("Email", "")),
        "Bundesland": data.get("Bundesland", ""),
        "Finanzierung": data.get("Welche Lösung passt am ehesten zu dir?", ""),
        "Adresse": data.get("Straße","") + ", " + data.get("Plz","") + " " + data.get("Ort",""),
        "Lead-ID": "",
        "Technologie": data.get("Thema",""),
        "Zeitraum": data.get("Umsetzung",data.get("Wie konkret ist dein Projekt aktuell?", "")),
        "Besitzverhältnis": "Eigentümer*in" if "ja" in (data.get("Eigentümer der Immobilie","")) else data.get("Eigentümer der Immobilie",""),
        "Stromspeicher": "Ja" if "Ja" in data.get("PF mit Speicher?","") else data.get("PF mit Speicher?",""),
        "Ergänzende Informationen": data.get("Warum interessierst du dich jetzt für eine PV-Lösung?", ""),
    }

# ============================================================
# ----------- MACHBARMACHER MSG EXTRACTION ----------
# ============================================================

# def extract_data_from_msg_machbarmacher(msg_path):
#     """Extract fields from MACHBARMACHER MSG file."""
#     msg = extract_msg.Message(msg_path)
#     body = msg.body
#     lines = [line.strip() for line in body.splitlines() if line.strip()]
#     #print("Answer" + lines.get("Wie konkret ist dein Projekt aktuell?",""))
#     data = {}
#     print(lines)

#     for index, line in enumerate(lines):
#         # Match "Key: Value"
#         match = re.match(r"([^:]+):\s*(.+)", line)
        
#         if match:
#             key = match.group(1).strip()
#             value = match.group(2).strip()
#             data[key] = value
        
#         # Special case: "Thema PV"
#         elif line.startswith("Thema"):
#             parts = line.split(" ", 1)
#             if len(parts) == 2:
#                 data["Thema"] = parts[1].strip()

#         elif index + 1 < len(lines):
#             # Detect question lines without colon
#             if line.endswith("?"):
#                 data[line.strip()] = lines[index + 1].strip()


#     # Extract name
#     full_name = data.get("Name", "")
#     vorname = full_name.split()[0] if full_name else ""
#     nachname = " ".join(full_name.split()[1:]) if len(full_name.split()) > 1 else ""

#     # Extract date/time
#     datum, uhrzeit = extract_datetime_from_zapier(body)

#     # Build return dict
#     return {
#         "Quelle": "MACHBARMACHER",
#         "Datum": datum,
#         "Uhrzeit": uhrzeit,
#         "Vorname": vorname,
#         "Nachname": nachname,
#         "Telefon": data.get("Tel", ""),
#         "E-Mail": clean_email(data.get("Email", "")),
#         "Bundesland": data.get("Bundesland", ""),
#         "Finanzierung": data.get("Welche Lösung passt am ehesten zu dir?", ""),
#         "Adresse": data.get("Straße","") + ", " + data.get("Plz","") + " " + data.get("Ort",""),
#         "Lead-ID": "",
#         "Technologie": data.get("Thema",""),
#         "Zeitraum": data.get("Umsetzung",data.get("Wie konkret ist dein Projekt aktuell?", "")),
#         "Besitzverhältnis": "Eigentümer*in" if "ja" in (data.get("Eigentümer der Immobilie","")) else data.get("Eigentümer der Immobilie",""),
#         "Stromspeicher": "Ja" if "Ja" in data.get("PF mit Speicher?","") else data.get("PF mit Speicher?",""),
#         "Ergänzende Informationen": data.get("Warum interessierst du dich jetzt für eine PV-Lösung?", ""),
#     }

# ============================================================
# ------------------ LOAD EXCEL ------------------------------
# ============================================================

wb = load_workbook(EXCEL_PATH)
ws = wb.active
headers = [cell.value for cell in ws[1]]

required_columns = ["Vorname", "Nachname", "Datum", "Uhrzeit","Lead-ID"]
for col in required_columns:
    if col not in headers:
        raise ValueError(f"Excel must contain column '{col}'")

vor_index = headers.index("Vorname")
nach_index = headers.index("Nachname")
date_index = headers.index("Datum")
time_index = headers.index("Uhrzeit")

existing_keys = set()

lead_index = headers.index("Lead-ID")

for row in ws.iter_rows(min_row=2, values_only=True): 
    lead_val = row[lead_index]
    if lead_val: 
        existing_keys.add(str(lead_val).strip())

# ---------- EXISTING KEYS MIT SHA FÜR PHOTOVOLTAIK ----------
for row in ws.iter_rows(min_row=2, values_only=True):
    nach = normalize(row[nach_index])
    excel_date = row[date_index]
    excel_time = row[time_index]

    if isinstance(excel_date, datetime):
        datum = excel_date.date().isoformat()
    elif hasattr(excel_date, "isoformat"):
        datum = excel_date.isoformat()
    else:
        datum = str(excel_date).strip()

    if isinstance(excel_time, datetime):
        uhrzeit = excel_time.time().strftime("%H:%M:%S")
    elif hasattr(excel_time, "strftime"):
        uhrzeit = excel_time.strftime("%H:%M:%S")
    else:
        uhrzeit = str(excel_time).strip()

    if nach and datum and uhrzeit:
        key_str = f"{nach}|{datum}|{uhrzeit}"
        key_hash = hashlib.sha256(key_str.encode("utf-8")).hexdigest()
        existing_keys.add(key_hash)

# ============================================================
# ---------------- PROCESS VOLTALUX PDFs ---------------------
# ============================================================

for filename in os.listdir(FOLDER_PATH_VOLTALUX):

    if not filename.lower().endswith(".pdf"):
        continue

    file_path = os.path.join(FOLDER_PATH_VOLTALUX, filename)

    try:
        with pdfplumber.open(file_path) as pdf:
            full_text = ""
            for page in pdf.pages:
                text = page.extract_text()
                if text:
                    full_text += text + "\n"

        data = extract_data_from_pdf(full_text)

        vor = normalize(data.get("Vorname"))
        nach = normalize(data.get("Nachname"))
        datum = str(data.get("Datum")).strip()
        uhrzeit = str(data.get("Uhrzeit")).strip()

        unique_key = data["Lead-ID"] if data["Lead-ID"] else data["E-Mail"]
        if not unique_key: 
            print(f"(Voltalux) No unique identifier in {filename} – skipping") 
            continue 
        if unique_key in existing_keys:
            print(f"(Voltalux) Duplicate found: {filename}") 
            continue

        new_row = [data.get(header, "") for header in headers]
        ws.append(new_row)
        existing_keys.add(unique_key)

        print(f"(Voltalux) Added: {filename}")

    except Exception as e:
        print(f"(Voltalux) Error processing {filename}: {e}")

# ============================================================
# -------- PROCESS PHOTOVOLTAIKANLAGE.AT MSG FILES ----------
# ============================================================

for filename in os.listdir(FOLDER_PATH_PHOTOVOLTAIKANLAGEAT):
    if not filename.lower().endswith(".msg"):
        continue
    file_path = os.path.join(FOLDER_PATH_PHOTOVOLTAIKANLAGEAT, filename)
    try:
        data = extract_data_from_msg_photovoltaikanlageat(file_path)
        if not data:
            continue
        vor = normalize(data.get("Vorname"))
        nach = normalize(data.get("Nachname"))
        datum = data.get("Datum")
        uhrzeit = data.get("Uhrzeit")
        if not (nach and datum and uhrzeit):
            print(f"(Photovoltaikanlage.at) Incomplete identity data in {filename}, skipping")
            continue

        # ------------------- SHA HASH CHECK -------------------
        key_str = f"{nach}|{datum}|{uhrzeit}"
        key_hash = hashlib.sha256(key_str.encode("utf-8")).hexdigest()
        if key_hash in existing_keys:
            print(f"(Photovoltaikanlage.at) Duplicate found: {filename}")
            continue

        new_row = [data.get(header,"") for header in headers]
        ws.append(new_row)
        existing_keys.add(key_hash)
        print(f"(Photovoltaikanlage.at) Added: {filename}")
    except Exception as e:
        print(f"(Photovoltaikanlage.at) Error processing {filename}: {e}")

# ============================================================
# -------- PROCESS PVALARM MSG FILES ----------
# ============================================================

for filename in os.listdir(FOLDER_PATH_PVALARM):
    if not filename.lower().endswith(".msg"):
        continue
    file_path = os.path.join(FOLDER_PATH_PVALARM, filename)
    try:
        data = extract_data_from_msg_pvalarm(file_path)
        if not data:
            continue
        vor = normalize(data.get("Vorname"))
        nach = normalize(data.get("Nachname"))
        datum = data.get("Datum")
        uhrzeit = data.get("Uhrzeit")

        if not (nach and datum and uhrzeit):
            print(f"(PVALARM) Incomplete identity data in {filename}, skipping")
            continue

        # ------------------- SHA HASH CHECK -------------------
        key_str = f"{nach}|{datum}|{uhrzeit}"
        key_hash = hashlib.sha256(key_str.encode("utf-8")).hexdigest()
        if key_hash in existing_keys:
            print(f"(PVALARM) Duplicate found: {filename}")
            continue

        new_row = [data.get(header,"") for header in headers]
        ws.append(new_row)
        existing_keys.add(key_hash)
        print(f"(PVALARM) Added: {filename}")
    except Exception as e:
        print(f"(PVALARM) Error processing {filename}: {e}")

# ============================================================
# -------- PROCESS MachbarMacher MSG FILES ----------
# ============================================================

# for filename in os.listdir(FOLDER_PATH_MACHBARMACHER):
#     if not filename.lower().endswith(".msg"):
#         continue
#     file_path = os.path.join(FOLDER_PATH_MACHBARMACHER, filename)
#     try:
#         data = extract_data_from_msg_machbarmacher(file_path)
#         if not data:
#             continue
#         vor = normalize(data.get("Vorname"))
#         nach = normalize(data.get("Nachname"))
#         datum = data.get("Datum")
#         uhrzeit = data.get("Uhrzeit")

#         if not (nach and datum and uhrzeit):
#             print(f"(MACHBARMACHER) Incomplete identity data in {filename}, skipping")
#             continue

#         # ------------------- SHA HASH CHECK -------------------
#         key_str = f"{nach}|{datum}|{uhrzeit}"
#         key_hash = hashlib.sha256(key_str.encode("utf-8")).hexdigest()
#         if key_hash in existing_keys:
#             print(f"(MACHBARMACHER) Duplicate found: {filename}")
#             continue

#         new_row = [data.get(header,"") for header in headers]
#         ws.append(new_row)
#         existing_keys.add(key_hash)
#         print(f"(MACHBARMACHER) Added: {filename}")
#     except Exception as e:
#         print(f"(MACHBARMACHER) Error processing {filename}: {e}")

# ============================================================
# ---------------- SAVE ONCE --------------------------------
# ============================================================

wb.save(EXCEL_PATH)
print("Processing complete.")
